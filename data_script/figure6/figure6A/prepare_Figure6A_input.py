from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SCRIPT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = SCRIPT_DIR.parents[1]
TABLE_S1 = PACKAGE_ROOT / "figure1" / "figure1cd" / "Supplement_Table1_Info_sample, related to figure1-revised.xlsx"
TABLE_S4 = PACKAGE_ROOT / "figure2" / "figure2b" / "Supplement_Table4_Normalized abundance, related to figure2-revised-S3-annotations.xlsx"
OUT = SCRIPT_DIR


def clean_text(x):
    return " ".join(str(x).strip().split())


def write_sheet(ws, frame):
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append(list(row))
    fill = PatternFill("solid", fgColor="E8EEF4")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 42)
        ws.column_dimensions[col[0].column_letter].width = width


def main():
    s4 = pd.read_excel(TABLE_S4, header=None)
    host_meta = pd.DataFrame(
        {
            "host_class": s4.iloc[1, 1:52].map(clean_text).to_numpy(),
            "host_order": s4.iloc[2, 1:52].map(clean_text).to_numpy(),
            "host_label": s4.iloc[3, 1:52].map(clean_text).to_numpy(),
        }
    )
    pathogen_meta = pd.DataFrame(
        {
            "pathogen": s4.iloc[4:, 0].map(clean_text).to_numpy(),
            "classification": s4.iloc[4:, 60].map(clean_text).to_numpy(),
            "taxonomy": s4.iloc[4:, 61].map(clean_text).to_numpy(),
            "novelty": s4.iloc[4:, 62].map(clean_text).to_numpy(),
            "multi_tissue_distribution": s4.iloc[4:, 63].map(clean_text).to_numpy(),
        }
    )
    rpm = s4.iloc[4:, 1:52].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    rpm.index = pathogen_meta["pathogen"]
    rpm.columns = host_meta["host_label"]

    raw_s1 = pd.read_excel(TABLE_S1, header=None)
    s1 = raw_s1.iloc[2:].copy()
    s1.columns = raw_s1.iloc[1].map(clean_text)
    tax = (
        s1[["class", "order / order-level lineage", "family", "genus", "species", "species abbreviation"]]
        .drop_duplicates()
        .rename(
            columns={
                "class": "host_class_s1",
                "order / order-level lineage": "host_order_s1",
                "species abbreviation": "host_label",
            }
        )
    )
    for col in tax.columns:
        tax[col] = tax[col].map(clean_text)
    host_meta = host_meta.merge(tax, on="host_label", how="left", validate="one_to_one")
    if host_meta["species"].isna().any():
        raise ValueError("Some Table S4 host labels could not be matched to Table S1")
    if not (host_meta["host_class"] == host_meta["host_class_s1"]).all():
        raise ValueError("Host class mismatch between Table S1 and Table S4")
    if not (host_meta["host_order"] == host_meta["host_order_s1"]).all():
        raise ValueError("Host order mismatch between Table S1 and Table S4")
    host_meta = host_meta[
        ["host_label", "species", "host_class", "host_order", "family", "genus"]
    ]

    class_to_type = {
        "RNA virus": "Virus",
        "DNA virus": "Virus",
        "Bacteria": "Bacteria",
        "Eukaryota": "Eukaryota",
    }
    pathogen_meta["node_type"] = pathogen_meta["classification"].map(class_to_type)
    if pathogen_meta["node_type"].isna().any():
        raise ValueError("Unrecognized pathogen classification")

    edge_rows = []
    for pathogen, row in rpm.iterrows():
        for host_label, value in row[row > 0].items():
            hm = host_meta.loc[host_meta["host_label"] == host_label].iloc[0]
            edge_rows.append(
                {
                    "source": f"host::{host_label}",
                    "target": f"pathogen::{pathogen}",
                    "host_label": host_label,
                    "host_species": hm["species"],
                    "host_class": hm["host_class"],
                    "host_order": hm["host_order"],
                    "pathogen": pathogen,
                    "RPM": float(value),
                    "log10_RPM_plus1": float(np.log10(value + 1.0)),
                }
            )
    edges = pd.DataFrame(edge_rows)

    positive_hosts = set(edges["host_label"])
    positive_pathogens = set(edges["pathogen"])
    host_out = host_meta[host_meta["host_label"].isin(positive_hosts)].copy()
    host_out["node_id"] = "host::" + host_out["host_label"]
    host_out["label"] = host_out["host_label"]
    host_out["node_type"] = "Host species"
    host_out["node_zone"] = "host_ring"
    host_out["connected_hosts"] = np.nan
    host_out["connected_orders"] = np.nan

    path_out = pathogen_meta[pathogen_meta["pathogen"].isin(positive_pathogens)].copy()
    summary = edges.groupby("pathogen").agg(
        connected_hosts=("host_label", "nunique"), connected_orders=("host_order", "nunique")
    )
    path_out = path_out.merge(summary, left_on="pathogen", right_index=True, validate="one_to_one")
    path_out["node_id"] = "pathogen::" + path_out["pathogen"]
    path_out["label"] = path_out["pathogen"]
    path_out["node_zone"] = np.where(path_out["connected_orders"] >= 2, "inner", "outer")
    for col in ["species", "host_class", "host_order", "family", "genus", "host_label"]:
        path_out[col] = ""

    node_cols = [
        "node_id", "label", "node_type", "node_zone", "host_label", "species",
        "host_class", "host_order", "family", "genus", "classification", "taxonomy",
        "novelty", "multi_tissue_distribution", "connected_hosts", "connected_orders",
    ]
    for col in ["classification", "taxonomy", "novelty", "multi_tissue_distribution"]:
        host_out[col] = ""
    nodes = pd.concat([host_out, path_out], ignore_index=True, sort=False)[node_cols]

    matrix_out = rpm.reset_index().rename(columns={"pathogen": "Pathogen"})
    readme = pd.DataFrame(
        {
            "item": [
                "Source Table S1", "Source Table S4", "Host inclusion", "Edge inclusion",
                "Edge width", "Host count", "Pathogen count", "Edge count",
            ],
            "value": [
                str(TABLE_S1), str(TABLE_S4), "Host species with at least one RPM > 0 edge",
                "All host-pathogen combinations with RPM > 0", "log10(RPM+1)",
                len(host_out), len(path_out), len(edges),
            ],
        }
    )

    nodes.to_csv(OUT / "Figure6A_node_table.tsv", sep="\t", index=False, encoding="utf-8-sig")
    edges.to_csv(OUT / "Figure6A_edge_table.tsv", sep="\t", index=False, encoding="utf-8-sig")

    wb = Workbook()
    wb.remove(wb.active)
    for name, frame in [
        ("host_metadata", host_meta),
        ("pathogen_metadata", pathogen_meta),
        ("abundance_RPM", matrix_out),
        ("node_table", nodes),
        ("edge_table", edges),
        ("README", readme),
    ]:
        write_sheet(wb.create_sheet(name), frame)
    wb.save(OUT / "Figure6A_input.xlsx")

    print(f"Hosts: {len(host_out)} / {len(host_meta)}")
    print(f"Pathogens: {len(path_out)} / {len(pathogen_meta)}")
    print(f"Edges: {len(edges)}")
    print(f"Inner pathogens: {(path_out['node_zone'] == 'inner').sum()}")
    print(f"Outer pathogens: {(path_out['node_zone'] == 'outer').sum()}")


if __name__ == "__main__":
    main()
